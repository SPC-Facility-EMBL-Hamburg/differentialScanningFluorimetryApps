import codecs
import re
import numpy  as np
import pandas as pd
from openpyxl import load_workbook
from xlrd     import open_workbook

from collections import Counter

from scipy.optimize     import least_squares

from constants import R_gas, temp_standard

def fit_line_robust(x,y):

    """
    Fit a line to the data using robust fitting
    Args:
        x (np.ndarray): x data
        y (np.ndarray): y data
    Returns:
        m (float): Slope of the fitted line
        b (float): Intercept of the fitted line
    """

    # convert x and y to numpy arrays, if they are lists
    if isinstance(x, list):
        x = np.array(x)
    if isinstance(y, list):
        y = np.array(y)

    def linear_model(z,params):
        slope,intercept = params
        return slope * z + intercept

    p0 = np.polyfit(x, y, 1)

    # Perform robust fitting
    res_robust = least_squares(
        lambda params: linear_model(x, params) - y,
        p0,
        loss='soft_l1',
        f_scale=0.1
    )

    m, b = res_robust.x

    return m, b


def fit_quadratic_robust(x,y):

    """
    Fit a quadratic equation to the data using robust fitting

    Args:
        x (np.ndarray): x data
        y (np.ndarray): y data
    Returns:
        a (float): Quadratic coefficient of the fitted line
        b (float): Linear coefficient of the fitted line
        c (float): Constant coefficient of the fitted line

    """

    # convert x and y to numpy arrays, if they are lists
    if isinstance(x, list):
        x = np.array(x)
    if isinstance(y, list):
        y = np.array(y)

    def model(x,params):
        a,b,c = params
        return a*np.square(x) + b*x + c

    p0 = np.polyfit(x, y, 2)

    # Perform robust fitting
    res_robust = least_squares(
        lambda params: model(x, params) - y,
        p0,
        loss='soft_l1',
        f_scale=0.1
    )

    a,b,c = res_robust.x

    return a,b,c

def temperature_to_kelvin(temp):
    """
    Convert temperature to Kelvin.

    Args:
        temp (array-like or float): Temperature value(s). If values are below 273.15
            they are assumed to be in Celsius and will be converted to Kelvin.

    Returns:
        numpy.ndarray or float: Temperatures in Kelvin with the same shape as the
        input.
    """

    if np.max(temp) < 273.15:
        # If the temperature is in Celsius
        return temp + 273.15
    else:
        # If the temperature is already in Kelvin
        return temp

def get_sheet_names_of_xlsx(filepath):

    """
    Get the sheet names of a .xls or .xlsx file without fully loading it.

    Args:
        filepath (str): Path to the Excel file.

    Returns:
        list: List of sheet names (strings) found in the file.
    """

    try:

        wb = load_workbook(filepath, read_only=True, keep_links=False)
        sheet_names = wb.sheetnames

    except:

        xls = open_workbook(filepath, on_demand=True)
        sheet_names =  xls.sheet_names()
 
    return sheet_names

def file_is_of_type_aunty(file_path):

    """
    Detect if file is an AUNTY xlsx file.

    The AUNTY format contains multiple sheets where the first column is
    temperatures and subsequent columns are fluorescence values. The first
    row contains the word 'wavelength' and the second row contains the word
    'temperature' in the first column.

    Args:
        file_path (str): Path to the .xls or .xlsx file to test.

    Returns:
        bool: True if the file matches the AUNTY format heuristic, False otherwise.
    """

    if not (file_path.endswith('.xls') or file_path.endswith('.xlsx')):
        return False

    sheet_names = get_sheet_names_of_xlsx(file_path)

    for sheet_name in sheet_names:

        # Load the data
        data = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None,skiprows=0)

        try:

            wavelength_cell = data.iloc[0, 1]
            temperature_cell = data.iloc[1, 0]
            corner_cell = data.iloc[0, 0]

            # Verify that the word wavelength is in the first row, second column
            condition1 = isinstance(wavelength_cell, str) and 'wavelength' in wavelength_cell.lower()

            # Verify that the word temperature is in the second row, first column
            condition2 = isinstance(temperature_cell, str) and 'temperature' in temperature_cell.lower()

            # Verify that the corner cell is empty (NaN)
            condition3 = np.isnan(corner_cell)

            if not (condition1 and condition2 and condition3):
                continue

            wavelengths = np.array(data.iloc[1, 1:]).astype(float)

            temperature = np.round(np.array(data.iloc[2:, 0]).astype(float),2)

            # Verify we have more than 5 temperatures
            condition4 = len(temperature) > 5

            # Verify we have more than 5 wavelengths
            condition5 = len(wavelengths) > 5

            if not (condition4 and condition5):
                continue

            return True

        except:

            return False

    return False

def file_is_of_type_uncle(xlsx_file):

    """
    Check if the file is an UNCLE-format file.

    Args:
        xlsx_file (str): Path to the Excel file to check.

    Returns:
        bool: True if the heuristic suggests this is an UNCLE file, False otherwise.
    """

    try:

        # Read the first sheet of the xlsx file
        data = pd.read_excel(xlsx_file,skiprows=1,nrows=5,header=None)

        # Extrac the first row
        row = data.iloc[0,1:]

        # concatenate the row into a string
        row_str = ' '.join([str(x) for x in row])

        # count the number of times we have word 'Time' and 'Temp'
        count_time = row_str.count('Time')
        count_temp = row_str.count('Temp')

        return count_time > 20 and count_temp == count_time

    except:
        return False

def detect_encoding(file_path):
    """
    Detect file encoding from a small candidate list.

    Args:
        file_path (str): Path to the file to probe.

    Returns:
        str: Detected encoding name or the string "Unknown encoding" if none matched.
    """
    encodings = ["utf-8", "latin1", "iso-8859-1", "cp1252"]
    for enc in encodings:
        try:
            with codecs.open(file_path, encoding=enc, errors="strict") as f:
                f.read()
            return enc
        except UnicodeDecodeError:
            continue
    return "Unknown encoding"

def is_blank(string):

    """
    Test whether a string is blank/empty or the literal "nan".

    Args:
        string (str or None): Input string to test.

    Returns:
        bool: True if the input is None, empty, whitespace-only, or the string "nan".
    """

    return (not (string and string.strip())) or string == "nan"

def find_indexes_of_non_signal_conditions(signal_data,conditions):

    """
    Find column indices that should be removed because they are non-signal.

    Args:
        signal_data (np.ndarray): 2D array (n_temps x n_conditions) of fluorescence
            values where NaN indicates missing/empty measurements.
        conditions (list or array-like): List of condition names (strings).

    Returns:
        list: Unique indices (ints) of columns that are non-signal (e.g. name
        contains derivative or 'S.D.' or column is entirely NaN).
    """

    # Find the indexes of conditions with derivative in it, or 'S.D.' in it
    idx_to_remove1 = [i for i, cond in enumerate(conditions) if 'derivative' in cond.lower() or 's.d.' in cond.lower()]

    # Find indexes empty signal_data columns
    idx_to_remove2 = [i for i in range(signal_data.shape[1]) if np.isnan(signal_data[:, i]).all()]

    # Find columns where all values are NaN
    nan_columns = np.all(np.isnan(signal_data), axis=0)

    # Get the column indices
    idx_to_remove3 = np.where(nan_columns)[0].tolist()

    # Find unique indexes to remove
    idx_to_remove = list(set(idx_to_remove1 + idx_to_remove2 + idx_to_remove3))

    return idx_to_remove

def find_repeated_words(string_lst):

    """
    Given a list of strings, find words that appear in every element (repeated).

    Args:
        string_lst (list of str): List of strings to analyse.

    Returns:
        list: List of words (str) that appear at least as many times as the
        number of elements in string_lst (i.e. repeated across entries).
    """

    # Repeated words
    words = " ".join(string_lst).split()

    # Count the occurrences of each word
    word_counts = Counter(words)

    # Find words that are repeated (appear, at least, as many times as elements in string_lst)
    repeated_words = [word for word, count in word_counts.items() if count >= len(string_lst)]

    return repeated_words

def remove_words_in_string(input_string,word_list):

    """
    Remove a list of words from an input string.

    Args:
        input_string (str): The input string.
        word_list (list of str): Words to remove from the string.

    Returns:
        str: The filtered string with the specified words removed.
    """

    # Split the string into words
    words = input_string.split()

    # Replace words from the list with an empty string
    filtered_words = [word for word in words if word not in word_list]

    # Join the words back into a single string
    output_string = ' '.join(filtered_words)

    return output_string

def subset_panta_data(data,columns_positions):

    """
    Subset PANTA-style data by extracting temperature and signal columns.

    Args:
        data (pandas.DataFrame): DataFrame with PANTA-style layout.
        columns_positions (list of int): Positions (1-based) identifying pairs of
            columns [temperature_col, signal_col] for each capillary/condition.

    Returns:
        tuple: (fluo, temp)
            fluo (np.ndarray): 2D array of fluorescence signals (n_temps x n_conditions).
            temp (np.ndarray): 1D array of temperatures (converted to Kelvin).
    """

    dfs    = [pd.DataFrame(np.array(data.iloc[:,[pos-1,pos]]),
        columns = ["temperature","signal"+str(i)]) for i,pos in enumerate(columns_positions)]

    # Case N = 1 (1 capillary)
    if len(dfs) == 1:

        mat   = dfs[0].to_numpy(dtype='float')

        fluo = mat[:,1:]         # You must use '1:' and not '1' to return a 2D array!
        temp = temperature_to_kelvin(mat[:,0])

        return fluo, temp

    # Combine dataframes so we can obtain a vector of temperatures and a matrix of fluorescence signal
    merged = pd.merge_asof(dfs[0].dropna(),dfs[1].dropna(),on="temperature", direction='nearest',allow_exact_matches=True)

    for df in dfs[2:]:
        
        merged = pd.merge_asof(merged,df.dropna(),on="temperature", direction='nearest',allow_exact_matches=True)       
    
    fluo   = np.array(merged.iloc[:, 1:]).astype('float')
    temp   = np.array(merged.iloc[:, 0]).astype('float')
    temp   = temperature_to_kelvin(temp)
    # Reduce data so we can plot and fit the data faster.

    while len(temp) > 700:
        fluo = fluo[::2]
        temp = temp[::2]

    return fluo, temp

def get_quantstudio_start_line(file_name):

    """
    Returns the number of the first line not starting with "*" + 1 for QuantStudio files.

    Args:
        file_name (str): Path to the QuantStudio text file.

    Returns:
        int: Line index (1-based) of the first non-comment line.
    """

    with codecs.open(file_name, 'r', encoding='utf-8',errors='ignore') as rf:
        ls       = rf.read().splitlines()
        splitted = [l.split() for l in ls]
        
        for i,s in enumerate(splitted):
            if len(s) > 5 and "*" not in s[0]:
                start_row = i+1
                break

    return start_row

def generate_qs3_df(df):

    """
    Generate a simple two-column DataFrame (temperature, signal) from a QuantStudio block.

    Args:
        df (pandas.DataFrame): Input block-like DataFrame where temperature and signal
            are in columns 3 and 4 (0-based indexing used here to match existing code).

    Returns:
        pandas.DataFrame: DataFrame with columns ["temperature", "signal"] sorted by temperature.
    """

    df = pd.DataFrame(np.array(df.iloc[:,[3,4]]),
        columns = ["temperature","signal"],dtype=str)

    df.replace(to_replace=',', value='',inplace=True,regex=True)

    df[["temperature", "signal"]] = df[["temperature", "signal"]].apply(pd.to_numeric)

    df.sort_values('temperature',inplace=True)

    return df

def generate_merged_quantstudio_df(data):

    """
    Merge multiple QuantStudio groups into temperature vector and fluorescence matrix.

    Args:
        data (pandas.DataFrame): Original QuantStudio DataFrame grouped by column 1.

    Returns:
        tuple: (fluo, temp)
            fluo (np.ndarray): 2D float array of fluorescence (n_temps x n_conditions).
            temp (np.ndarray): 1D float array of temperatures (in Kelvin).
    """

    dfs = [generate_qs3_df(x) for _, x in data.groupby(1)]
    dfs = [df.rename(columns={"signal": 'signal'+str(i)}) for i,df in enumerate(dfs)]

    # Combine dataframes so we can obtain a vector of temperatures and a matrix of fluorescence signal
    merged = pd.merge_asof(dfs[0].dropna(),dfs[1].dropna(),on="temperature", direction='nearest',allow_exact_matches=True)

    for df in dfs[2:]:
        
        merged = pd.merge_asof(merged,df.dropna(),on="temperature", direction='nearest',allow_exact_matches=True)       

    fluo   = np.array(merged.iloc[:, 1:]).astype('float')
    temp   = np.array(merged.iloc[:, 0]).astype('float')
    temp   = temperature_to_kelvin(temp)

    # Reduce data so we can plot and fit the data faster.

    n_conditions = fluo.shape[1]

    if n_conditions > 192:
        max_points = 150

    elif n_conditions > 96:
        max_points = 300

    else:
        max_points = 700

    while len(temp) > max_points:
        fluo = fluo[::2]
        temp = temp[::2]

    return fluo, temp

def detect_txt_file_type(file):

    """
    Decide if a text file was generated by an Agilent MX3005P or a QuantStudio instrument.

    Args:
        file (str): Path to the text file to inspect.

    Returns:
        str or None: 'MX3005P', 'QuantStudio', or None if undetermined.
    """

    with codecs.open(file, 'r', encoding='utf-8',errors='ignore') as rf:
        ls       = rf.read().splitlines()

        for line in ls:
            if line.startswith('Segment') and 'Well' in line:
                return 'MX3005P'

            splitted_line = line.split()
            if 'Well' in splitted_line and 'Target' in splitted_line and 'Reading' in splitted_line:
                return 'QuantStudio'

    return None

def load_layout(xls_file):

    """
    Load a layout Excel file and return a cleaned list of condition strings.

    Args:
        xls_file (str): Path to the Excel file containing layout information.

    Returns:
        numpy.ndarray: 1D array of cleaned condition strings (one per row).
    """

    xls = pd.ExcelFile(xls_file)
    dat = pd.read_excel(xls)

    conditions =  []

    for row in range(dat.shape[0]):

        condition = dat.iloc[row,1:].values
        condition = ' '.join([str(x) for x in condition])
        condition = condition.replace('nan','')
        condition = " ".join(condition.split())

        conditions.append(condition)

    return np.array(conditions)

def filter_fluo_by_temp(np_fluo,np_temp,min_temp,max_temp):

    """
    Filter the fluorescence matrix using a temperature range.

    Args:
        np_fluo (np.ndarray): 2D array (n_temps x n_conditions) or 1D array of fluorescence.
        np_temp (np.ndarray): 1D array of temperatures with length n_temps.
        min_temp (float): Minimum temperature to keep (inclusive).
        max_temp (float): Maximum temperature to keep (inclusive).

    Returns:
        np.ndarray: Filtered fluorescence matrix (rows where temperature within bounds).
    """

    shape = np_fluo.shape
    # Convert 1D array to a 2D numpy array of n rows and 1 column
    if len(shape) == 1:
        np_fluo = np.reshape(np_fluo, (-1, 1))

    tot = np_fluo.shape[1]

    np_tog = np.column_stack((np_fluo, np_temp))
    np_tog = np_tog[np_tog[:, tot] >= min_temp]
    np_tog = np_tog[np_tog[:, tot] <= max_temp]
    np_tog = np.delete(np_tog, tot, 1)

    return np_tog

def filter_temp_by_temp(np_temp,min_temp,max_temp):

    """
    Filter the temperature vector using a temperature range.

    Args:
        np_temp (np.ndarray): 1D array of temperatures.
        min_temp (float): Minimum temperature to keep (inclusive).
        max_temp (float): Maximum temperature to keep (inclusive).

    Returns:
        np.ndarray: Filtered temperature vector.
    """

    np_temp_filter = np_temp[np_temp >= min_temp]
    np_temp_filter = np_temp_filter[np_temp_filter <= max_temp]

    return np_temp_filter

def median_filter_from_fluo_and_temp_vectors(fluo_vec,temp_vec,rolling_window):

    """

    Compute the median filter of the fluorescence vector using a temperature rolling window.

    The function converts temperatures to integer seconds (scaled) and uses a pandas
    rolling median with a time-based window.

    Args:
        fluo_vec (array-like): 1D fluorescence values.
        temp_vec (array-like): 1D temperature values (same length as fluo_vec).
        rolling_window (float): Rolling window in Celsius degrees.

    Returns:
        numpy.ndarray: 1D array with the median-filtered fluorescence values.

    """

    scaling_factor = 10000

    temp_vec     =  np.multiply(temp_vec,scaling_factor).astype(int) 
    series       =  pd.Series(fluo_vec,index=temp_vec,dtype=float)
    series.index =  pd.to_datetime(series.index,unit='s')

    roll_window  = str(int(rolling_window*scaling_factor))+"s"

    fluo_df_median_filt = series.rolling(roll_window).median().to_numpy()

    return fluo_df_median_filt

def get_temp_at_maximum_of_derivative(temps,fluo_derivative):

    """
    Get temperature(s) corresponding to the maximum of the derivative for each column.

    Args:
        temps (array-like): 1D array of temperatures (length m).
        fluo_derivative (np.ndarray): 2D array (m x n) of derivative values.

    Returns:
        numpy.ndarray: 1D array (length n) of temperatures at which derivative is maximal per column.
    """

    tms_derivative = np.take(temps, np.argmax(fluo_derivative,axis=0))
    return tms_derivative

def get_temp_at_minimum_of_derivative(temps,fluo_derivative):

    """
    Get temperature(s) corresponding to the minimum of the derivative for each column.

    Args:
        temps (array-like): 1D array of temperatures (length m).
        fluo_derivative (np.ndarray): 2D array (m x n) of derivative values.

    Returns:
        numpy.ndarray: 1D array (length n) of temperatures at which derivative is minimal per column.
    """

    tms_derivative = np.take(temps, np.argmin(fluo_derivative,axis=0))
    return tms_derivative

def estimate_initial_tm_from_baseline_fitting(bUs,bNs,kNs,kUs,temps,fit_length,fluo_derivative):


    """
    Use a heuristic algorithm to get an initial estimate of the Tms.

    Args:
        bUs (array-like): Intercepts for unfolded baseline (length K).
        bNs (array-like): Intercepts for native baseline (length K).
        kNs (array-like): Slopes for native baseline (length K).
        kUs (array-like): Slopes for unfolded baseline (length K).
        temps (array-like): 1D temperature vector.
        fit_length (float): Temperature range used in fitting baseline heuristics.
        fluo_derivative (np.ndarray): 2D array (len(temps) x K) with derivative values.

    Returns:
        numpy.ndarray: 1D array of estimated Tm values (length K).
    """

    dintersect = - (bUs - bNs) / (kNs - kUs)
    
    tmin = min(temps)
    tmax = max(temps)
    tmid = (tmin + tmax) / 2

    c1 = dintersect > (tmin+fit_length)
    c2 = dintersect < (tmax+fit_length)
    c3 = np.logical_and(c1,c2)

    tms_case1 = np.where(c3,tmid,0)

    c4 = np.logical_not(c3)

    b_diff = (kUs - kNs ) * tmid + (bUs - bNs)

    c5 = b_diff > 0
    c6 = np.logical_and(c4,c5)

    tms_case2 = get_temp_at_maximum_of_derivative(temps,fluo_derivative) * c6
    
    c7 = np.logical_not(c5)
    c8 = np.logical_and(c7,c4)

    tms_case3 = np.take(temps, np.argmin(fluo_derivative,axis=0)) * c8

    tms = tms_case1 + tms_case2 + tms_case3

    return tms

def get_two_state_deltaG(dHms, tms, cp):

    """

    Obtain Gibbs free energy of unfolding at a standard temperature extrapolated using dCp.

    Args:
        dHms (array-like): Enthalpy changes at Tm (same shape as tms).
        tms (array-like): Melting temperatures (Kelvin).
        cp (float or array-like): Heat capacity change (dCp).

    Returns:
        tuple: (dG_std, dCp_component)
            dG_std (numpy.ndarray): Gibbs free energy at temp_standard.
            dCp_component (numpy.ndarray): The dCp contribution factor used in extrapolation.
    """
    dHms = np.array(dHms)
    tms  = np.array(tms)

    dG_std = dHms * (1 - temp_standard / tms) - cp * (tms - temp_standard + temp_standard * np.log(temp_standard / tms))

    # Calculate contribution of dCp to the real dG 
        
    dCp_component = temp_standard - tms - temp_standard * np.log(temp_standard / tms)

    return dG_std, dCp_component

def get_two_state_tonset(dHms, tms, onset_threshold=0.01):

    """

    Compute onset temperature based on fitted dHm and Tm using an onset threshold.

    Args:
        dHms (array-like): Enthalpy changes at Tm (same shape as tms).
        tms (array-like): Melting temperatures (Kelvin).
        onset_threshold (float): Fraction unfolded at onset (default 0.01).

    Returns:
        numpy.ndarray: Estimated onset temperatures (same shape as tms).
    """

    dHms,    tms        = np.array(dHms),    np.array(tms) 

    T_onset = 1 / (1 / tms - R_gas / dHms * np.log(onset_threshold / (1 - onset_threshold)))

    return T_onset

def get_empirical_two_state_score(Tm, T_onset):

    """

    Compute an empirical two-state quality score as Euclidean distance of (Tm, T_onset).

    Args:
        Tm (array-like): Melting temperatures (Kelvin or same units as T_onset).
        T_onset (array-like): Onset temperatures (same shape as Tm).

    Returns:
        numpy.ndarray: Euclidean distance per element: sqrt(Tm**2 + T_onset**2).
    """

    return np.sqrt(np.array(Tm)**2 + np.array(T_onset)**2)

def get_eq_three_state_combined_deltaG(dHm1_fit, T1_fit, dHm2_fit, T2_fit):

    """
    Compute combined deltaG for an equilibrium three-state model extrapolated to standard temperature.

    Args:
        dHm1_fit (float) : Enthalpy of first transition.
        T1_fit   (float) : Transition temperature (Kelvin) of first transition.
        dHm2_fit (float) : Enthalpy of second transition.
        T2_fit   (float) : Transition temperature (Kelvin) of second transition.

    Returns:
        numpy.ndarray: Combined deltaG at temp_standard.
    """

    dHm1_fit,   dHm2_fit    = np.array(dHm1_fit), np.array(dHm2_fit)
    T1_fit,     T2_fit      = np.array(T1_fit),   np.array(T2_fit)

    dG_comb_std = dHm1_fit * (1 - temp_standard / T1_fit) + dHm2_fit * (1 - temp_standard / T2_fit)

    return dG_comb_std

def get_empirical_three_state_score(T_onset1_fit, T1_fit, T_onset2_fit, T2_fit):

    """
    Args:
        T_onset1_fit (float): Onset temperature of first transition.
        T1_fit (float): Transition temperature of first transition.
        T_onset2_fit (float): Onset temperature of second transition.
        T2_fit (float): Transition temperature of second transition.

    Returns:
        float: Distance metric

    Notes:
        The metric is the distance from the origin (0,0) to the point (T1_fit,Tonset1) +
        the distance from the origin (0,0) to the point (T2_fit,Tonset2)

    """

    T_onset1_fit,   T_onset2_fit    = np.array(T_onset1_fit),   np.array(T_onset2_fit)
    T1_fit,         T2_fit          = np.array(T1_fit),         np.array(T2_fit)

    T_eucl_comb = np.sqrt(T1_fit**2 + T_onset1_fit**2) + np.sqrt(T2_fit**2 + T_onset2_fit**2)

    return T_eucl_comb

def arrhenius(T, Tf, Ea):
    """
    Arrhenius equation: dependence of reaction rate constant on temperature.

    Args:
        T (float or array-like): Temperature(s) at which to evaluate the function.
        Tf (float or array-like): Reference temperature(s) where k = 1 in this parametrisation.
        Ea (float or array-like): Activation energy (same units as R_gas*T).

    Returns:
        numpy.ndarray or float: exp(-Ea/R * (1/T - 1/Tf)).
    """
    return np.exp(-Ea / R_gas * (1 / T - 1 / Tf))

def get_irrev_two_state_pkd(Tf, Ea):

    """
    Compute irreversible two-state pKd-like value from Arrhenius parameters.

    Args:
        Tf (float or array-like): Reference temperature(s) for which k=1.
        Ea (float or array-like): Activation energy (same shape as Tf).

    Returns:
        numpy.ndarray or float: -log10(k(temp_standard)) where k is computed from arrhenius().
    """

    Tf = np.array(Tf)
    Ea = np.array(Ea)

    return -np.log10(arrhenius(temp_standard, Tf, Ea))

def get_barycenter(intensities,wavelengths):

    """
    Compute a (intensity-weighted) barycenter of wavelengths.

    Args:
        intensities (array-like): Intensity values (weights).
        wavelengths (array-like): Wavelength values (same shape as intensities) or matrix.

    Returns:
        numpy.ndarray or float: Weighted barycenter value(s).
    """

    barycenter = np.sum(wavelengths * intensities) / np.sum(intensities)

    return barycenter

def detect_file_type(file):

    """
    Detect the type of a given data file based on its extension and content.
    Args:
        file (str): Path to the data file.
    Returns:
        str or None: Detected file type (e.g., 'csv', 'thermofluor', 'panta', 'tycho', 'aunty', 'uncle', 'prometheus', 'supr', 'MX3005P', 'QuantStudio') or None if undetermined.
    """

    file_type = None

    if file.endswith(".txt"):

        file_type = detect_txt_file_type(file)

    if file.endswith(".csv"):

        file_type = "csv"

    if file.endswith(".supr"):

        file_type = "supr"

    if file.endswith(".xlsx") or file.endswith(".xls"):

        sheet_names = get_sheet_names_of_xlsx(file)

        if "RFU" in sheet_names:

            file_type = "thermofluor"

        elif "Data Export" in sheet_names or "melting-scan" in sheet_names:

            file_type = "panta"

        elif "Profiles_raw" in sheet_names:

            file_type = "tycho"

        elif file_is_of_type_aunty(file):

            file_type = "aunty"

        elif file_is_of_type_uncle(file):

            file_type = "uncle"

        else:

            file_type = "prometheus"

    return file_type

def generate_2D_signal_matrix(condition_id,signal_data_dictionary,selected_rows):

    """
    Args:
        condition_id (integer): Condition identifier.
        signal_data_dictionary (dict): Dictionary containing signal data (keys: wavelengths, values: signal values).
        selected_rows (list): List of wavelength indices to keep.
    Returns:
        numpy.ndarray: 2D array of signal values (n_temps x n_wavelengths).
    """

    signals_i = [signal[:,condition_id] for signal in signal_data_dictionary.values()]
    signals_2D = np.array(signals_i)[selected_rows]

    return signals_2D

def string_to_float(s):

    """
    Args:
        s (str): String to convert.
    Returns:
        float: Converted float value.
    """

    cleaned = re.sub(r'[^0-9.\-]', '', s)
    try:
        cleaned = float(cleaned)
    except ValueError:
        cleaned = 0.0
    return cleaned

def find_closest_signal(selected,options):

    """

    Args:
        selected (str): Selected condition string.
        options (list): List of available signals

    Returns:

    """

    # Remove all non-numeric characters from selected and from the options
    options_float = [string_to_float(s) for s in options]
    selected_float = string_to_float(selected)

    idx = np.argmin(np.abs(np.array(options_float)-selected_float))

    return options[idx]